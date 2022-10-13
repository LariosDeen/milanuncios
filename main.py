import os
from datetime import date

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from requests_html import HTMLSession

START_PRICE = 600


class Blocked(Exception):
    pass


def get_number(lnk: str):
    """Returns an integer number of properties on the web page by reference lnk."""
    r = session.get(lnk)
    try:
        about = r.html.find('.ma-ContentListingSummary-label', first=True)
    except Exception:
        return 0

    try:
        target_str = about.text
    except Exception:
        raise Blocked('Your enter is blocked. Wait 60 minutes or change IP')

    if ' anuncios' in target_str:
        target_number = int(target_str[:target_str.find(' anuncios')])
        return target_number
    else:
        return 0


def prises_list(start_price: int):
    """List of twains of price forks."""
    lst = []
    for i in range(0, 301, 100):
        twain = (str(start_price + i), str(start_price + i + 99))
        lst.append(twain)
    return lst


def visualization(count: int):
    """Visualization of the script process in terminal."""
    zero = lambda x: '' if x >= 10 else '0'
    num_dict = {1: 'st', 2: 'nd', 3: 'rd'}
    empty_space = 10 - count
    if count < 4:
        suffix = num_dict[count]
    else:
        suffix = 'th'
    return f'Parsing of {zero(count)}{count}-{suffix} page' + ' ' * empty_space + ' ...'


def excel_writing(filename: str):
    """Fills the excel file"""
    # working with file filename ('Example.xlsm')
    book = load_workbook(filename=filename, read_only=False, keep_vba=True)
    sheet = book.active
    # next_row - a string to be added to the end of an existing table of the file filename
    next_row = sheet.max_row + 1
    # variable for formatting the added string to the current sheet of the file filename
    medium_border = Border(left=Side(style='medium'),
                           right=Side(style='medium'),
                           top=Side(style='medium'),
                           bottom=Side(style='medium'))
    # if the date of the first cell of the last row of the active sheet of the file
    # filename is different from the current date
    if str(sheet[sheet.max_row][0].value)[0:10] != str(date.today()):
        # add the current date to the beginning of the list final_list
        final_list.insert(0, date.today())
        # add a new line to the active sheet of the filename file after the last one
        sheet.append(final_list)
        # set the formatting of the added line in the active sheet of the filename file
        sheet.row_dimensions[next_row].height = 23.6
        sheet[next_row][0].fill = PatternFill(fill_type='solid', start_color='FFB6FCC5')
        sheet[next_row][9].fill = PatternFill(fill_type='solid', start_color='FFF2F2F2')
        sheet[next_row][10].fill = PatternFill(fill_type='solid', start_color='FFF2F2F2')
        sheet[next_row][0].font = Font(size=10)
        for i in range(1, 12):
            sheet.cell(row=next_row, column=i).border = medium_border
            sheet.cell(row=next_row, column=i).alignment = Alignment(
                horizontal='center', vertical='center')
        for i in range(1, 11):
            sheet[next_row][i].font = Font(size=18)
        # set the numeric date format (column A) as in the previous cell on top
        sheet[next_row][0].number_format = sheet[next_row - 1][0].number_format
    # save the changes to the file
    book.save(filename)


towns = ('malaga', 'torremolinos')
final_list = []
session = HTMLSession()
counter = 0

if __name__ == '__main__':
    for town in towns:
        for prise_from, prise_to in prises_list(START_PRICE):
            link = f'https://www.milanuncios.com/alquiler-de-pisos-en-{town}-malaga/?' \
                   f'desde={prise_from}&hasta={prise_to}&demanda=n&banosd=2&dormd=3'
            number = get_number(link)
            final_list.append(number)
            counter += 1
            print(visualization(counter))

    for town in towns:
        link = f'https://www.milanuncios.com/alquiler-de-pisos-en-{town}-malaga/?demanda=n&banosd=1&dormd=4'
        number = get_number(link)
        final_list.append(number)
        counter += 1
        print(visualization(counter))

    print(final_list)
    excel_writing('Estadistica.xlsm')
    # open the file Estadistica.xlsm for visual viewing
    os.startfile('Estadistica.xlsm')
